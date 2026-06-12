# Step 4: Auto-generate Excel report with pivot tables and charts
# Usage: python scripts/generate_excel.py
# Output: excel/job_market_report.xlsx

import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

os.makedirs(r"C:\Users\simmi\OneDrive\Desktop\job-market-intelligence\excel", exist_ok=True)

# LOAD DATA
df = pd.read_csv(r"C:\Users\simmi\OneDrive\Desktop\job-market-intelligence\data\processed\job_market_combined.csv")
df_salary  = df[df["source"] == "salary_ds"].copy()
df_fresher = df[df["source"] == "fresher_ds"].copy()
df_fresher["hired"] = pd.to_numeric(df_fresher["hired"], errors="coerce").fillna(0).astype(int)

print("Data loaded. Building Excel report...")

wb = Workbook()


# STYLING HELPERS
TEAL       = "0F6E56"
TEAL_LIGHT = "E1F5EE"
PURPLE     = "534AB7"
PURPLE_LT  = "EEEDFE"
GRAY_HDR   = "444441"
GRAY_LIGHT = "F1EFE8"
WHITE      = "FFFFFF"
BORDER_COLOR = "D3D1C7"

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

def write_header_row(ws, row_num, headers, fill_color=TEAL, font_color=WHITE):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.font      = Font(bold=True, color=font_color, size=11)
        cell.fill      = hdr_fill(fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()

def write_data_row(ws, row_num, values, fill_hex=None):
    fill = hdr_fill(fill_hex) if fill_hex else None
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=v)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()
        if fill:
            cell.fill = fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_cell(ws, row, col, text, color=TEAL):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, size=13, color=color)
    c.alignment = Alignment(horizontal="left", vertical="center")

# SHEET 1: SUMMARY DASHBOARD
ws1 = wb.active
ws1.title = "Dashboard"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40
ws1.merge_cells("A1:G1")
title = ws1["A1"]
title.value     = "Job Market Intelligence - India"
title.font      = Font(bold=True, size=16, color=WHITE)
title.fill      = hdr_fill(TEAL)
title.alignment = Alignment(horizontal="center", vertical="center")

# KPI cards (row 3)
kpis = [
    ("Total listings",    len(df_salary)),
    ("Fresher applications", len(df_fresher)),
    ("Unique cities",     df_salary["city"].nunique()),
    ("Unique roles",      df_salary["job_role"].nunique()),
    ("Median salary ₹",   int(df_salary["salary_inr"].median())),
    ("Overall hire rate", f"{100*df_fresher['hired'].mean():.1f}%"),
]
ws1.row_dimensions[3].height = 15
ws1.row_dimensions[4].height = 35
ws1.row_dimensions[5].height = 30
for col_idx, (label, value) in enumerate(kpis, 1):
    lc = ws1.cell(row=4, column=col_idx, value=label)
    lc.font      = Font(bold=True, size=10, color=GRAY_HDR)
    lc.alignment = Alignment(horizontal="center")
    lc.fill      = hdr_fill(TEAL_LIGHT)
    vc = ws1.cell(row=5, column=col_idx, value=value)
    vc.font      = Font(bold=True, size=14, color=TEAL)
    vc.alignment = Alignment(horizontal="center")
    vc.fill      = hdr_fill(WHITE)
    for r in [4, 5]:
        ws1.cell(r, col_idx).border = thin_border()

set_col_widths(ws1, [22, 22, 18, 18, 18, 18, 18])
ws1["A7"].value = "Refer to the sheets below for full pivot tables and charts."
ws1["A7"].font  = Font(italic=True, color="888780", size=10)


# SHEET 2: SALARY BY CITY (pivot)
ws2 = wb.create_sheet("Salary by City")
ws2.sheet_view.showGridLines = False

salary_pivot = (
    df_salary.groupby("city")["salary_inr"]
    .agg(avg_salary="mean", median_salary="median",
         min_salary="min", max_salary="max", listings="count")
    .query("listings >= 5")
    .sort_values("avg_salary", ascending=False)
    .head(20)
    .reset_index()
)
salary_pivot["avg_salary"]    = salary_pivot["avg_salary"].round(0).astype(int)
salary_pivot["median_salary"] = salary_pivot["median_salary"].round(0).astype(int)
salary_pivot["min_salary"]    = salary_pivot["min_salary"].round(0).astype(int)
salary_pivot["max_salary"]    = salary_pivot["max_salary"].round(0).astype(int)

title_cell(ws2, 1, 1, "Average Salary by City (Top 20)")
write_header_row(ws2, 3, ["City", "Avg Salary (₹)", "Median (₹)", "Min (₹)", "Max (₹)", "Listings"])
for i, row in enumerate(salary_pivot.itertuples(), 4):
    fill = TEAL_LIGHT if i % 2 == 0 else WHITE
    write_data_row(ws2, i, [row.city, row.avg_salary, row.median_salary,
                             row.min_salary, row.max_salary, row.listings], fill)
set_col_widths(ws2, [22, 16, 16, 14, 14, 12])

# Add bar chart
chart2 = BarChart()
chart2.type  = "bar"
chart2.title = "Top 20 Cities - Average Salary"
chart2.y_axis.title = "City"
chart2.x_axis.title = "Avg Salary (₹)"
chart2.height = 14
chart2.width  = 22
data_ref    = Reference(ws2, min_col=2, max_col=2, min_row=3, max_row=3+len(salary_pivot))
labels_ref  = Reference(ws2, min_col=1, max_col=1, min_row=4, max_row=3+len(salary_pivot))
chart2.add_data(data_ref, titles_from_data=True)
chart2.set_categories(labels_ref)
ws2.add_chart(chart2, "H3")



# SHEET 3: HIRING ANALYSIS (pivot)
ws3 = wb.create_sheet("Hiring Analysis")
ws3.sheet_view.showGridLines = False

# Platform hire rate
platform_pivot = (
    df_fresher.groupby("platform")
    .agg(total=("hired","count"), hired=("hired","sum"))
    .assign(hire_rate=lambda x: (100 * x["hired"] / x["total"]).round(1))
    .sort_values("hire_rate", ascending=False)
    .reset_index()
)
title_cell(ws3, 1, 1, "Hire Rate by Platform")
write_header_row(ws3, 3, ["Platform", "Total Applications", "Hired", "Hire Rate (%)"],
                 fill_color=PURPLE, font_color=WHITE)
for i, row in enumerate(platform_pivot.itertuples(), 4):
    fill = PURPLE_LT if i % 2 == 0 else WHITE
    write_data_row(ws3, i, [row.platform, row.total, row.hired, row.hire_rate], fill)

# CGPA pivot (below platform table, with gap)
gap_row = 4 + len(platform_pivot) + 2
title_cell(ws3, gap_row, 1, "Hire Rate by CGPA Band")
write_header_row(ws3, gap_row+2, ["CGPA Band", "Total", "Hired", "Hire Rate (%)"],
                 fill_color=PURPLE, font_color=WHITE)
df_fresher_cgpa = df_fresher.copy()
df_fresher_cgpa["cgpa"] = pd.to_numeric(df_fresher_cgpa["cgpa"], errors="coerce")
df_fresher_cgpa["cgpa_band"] = pd.cut(
    df_fresher_cgpa["cgpa"], bins=[0, 6.5, 7.5, 8.5, 10],
    labels=["Below 6.5", "6.5–7.4", "7.5–8.4", "8.5+"]
)
cgpa_pivot = (
    df_fresher_cgpa.groupby("cgpa_band", observed=True)
    .agg(total=("hired","count"), hired=("hired","sum"))
    .assign(hire_rate=lambda x: (100 * x["hired"] / x["total"]).round(1))
    .reset_index()
)
for i, row in enumerate(cgpa_pivot.itertuples(), gap_row+3):
    fill = PURPLE_LT if i % 2 == 0 else WHITE
    write_data_row(ws3, i, [str(row.cgpa_band), row.total, row.hired, row.hire_rate], fill)

set_col_widths(ws3, [22, 20, 12, 16])


# SHEET 4: SKILLS ANALYSIS

ws4 = wb.create_sheet("Skills Analysis")
ws4.sheet_view.showGridLines = False

all_skills = (
    df_fresher["skills"].dropna()
    .str.split(",").explode()
    .str.strip().str.lower()
    .value_counts()
    .head(25)
    .reset_index()
)
all_skills.columns = ["Skill", "Frequency"]

hired_skills = (
    df_fresher[df_fresher["hired"] == 1]["skills"].dropna()
    .str.split(",").explode()
    .str.strip().str.lower()
    .value_counts()
    .head(25)
    .reset_index()
)
hired_skills.columns = ["Skill", "Hired_Frequency"]

title_cell(ws4, 1, 1, "Top Skills - All Applicants")
write_header_row(ws4, 3, ["Skill", "Frequency (All)"])
for i, row in enumerate(all_skills.itertuples(), 4):
    fill = TEAL_LIGHT if i % 2 == 0 else WHITE
    write_data_row(ws4, i, [row.Skill, row.Frequency], fill)

title_cell(ws4, 1, 4, "Top Skills - Hired Candidates", color=PURPLE)
write_header_row(ws4, 3, ["Skill", "Frequency (Hired)"], fill_color=PURPLE,
                 font_color=WHITE)
for i, row in enumerate(hired_skills.itertuples(), 4):
    fill = PURPLE_LT if i % 2 == 0 else WHITE
    c1 = ws4.cell(row=i, column=4, value=row.Skill)
    c2 = ws4.cell(row=i, column=5, value=row.Hired_Frequency)
    for c in [c1, c2]:
        c.alignment = Alignment(horizontal="center")
        c.border    = thin_border()
        if fill: c.fill = hdr_fill(fill)

set_col_widths(ws4, [22, 18, 5, 22, 20])


# SAVE
output_path = r"C:\Users\simmi\OneDrive\Desktop\job-market-intelligence\excel\job_market_report.xlsx"
wb.save(output_path)
print(f"\n[Done] Excel report saved: {output_path}")
print("  Sheets created:")
print("    1. Dashboard     - KPI summary")
print("    2. Salary by City - pivot + bar chart")
print("    3. Hiring Analysis - platform + CGPA pivots")
print("    4. Skills Analysis - top skills comparison")
print("\nNext: Open in Excel, add slicers manually via Insert -> Slicer")
print("Then: Build Power BI dashboard (Step 5)")